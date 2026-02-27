"""
Excel → requirements JSON parser.

Parses a customer requirements Excel file (e.g. "Customer_Metrics.xlsx")
into the canonical JSON structure expected by the builder:

    [
      {
        "name": "Domain 1 of 6 · Application & Model Inference Telemetry",
        "description": "22 requirements · Cost allocation, ...",
        "reqs": [
          {"requirement": "...", "description": "...", "status": "✅ Now", "signal": "TRACE"},
          ...
        ]
      },
      ...
    ]

Supports two common formats:
  1. Multi-sheet: one sheet per domain, rows are requirements.
  2. Single-sheet: one sheet with a "Domain" column to group by.
"""
import json
import os
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None  # Handled at call time


# ─────────────────────────────────────────────────────────────────────────────
# Column name aliases (case-insensitive matching)
# ─────────────────────────────────────────────────────────────────────────────
_ALIASES = {
    "requirement": ["requirement", "req", "requirement name", "name", "capability"],
    "description": ["description", "desc", "detail", "details", "notes", "note"],
    "status":      ["status", "coverage", "availability", "state", "response"],
    "signal":      ["signal", "signal type", "telemetry", "data type", "type"],
    "domain":      ["domain", "domain name", "category", "group", "area"],
}


def _match_col(header: str) -> Optional[str]:
    """Map a raw column header to a canonical key, or None."""
    h = header.strip().lower()
    for key, aliases in _ALIASES.items():
        if h in aliases:
            return key
    return None


def _normalise_status(val: str) -> str:
    """Ensure status values include the standard emoji prefixes."""
    v = val.strip()
    lo = v.lower()
    if "now" in lo or "available" in lo or "yes" in lo:
        return f"✅ {v}" if "✅" not in v else v
    if "partial" in lo:
        return f"⚡ {v}" if "⚡" not in v else v
    if "roadmap" in lo or "planned" in lo or "future" in lo:
        return f"🗺 {v}" if "🗺" not in v else v
    return v


# ─────────────────────────────────────────────────────────────────────────────
# Parse a single sheet into a list of requirement dicts
# ─────────────────────────────────────────────────────────────────────────────
def _parse_sheet(ws) -> list[dict]:
    """Read requirement rows from a worksheet. Returns list of req dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first row with ≥2 recognised columns)
    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        mapping = {}
        for c, cell_val in enumerate(row):
            if cell_val is None:
                continue
            key = _match_col(str(cell_val))
            if key:
                mapping[key] = c
        if len(mapping) >= 2:
            header_idx = i
            col_map = mapping
            break

    if header_idx is None:
        return []

    reqs = []
    for row in rows[header_idx + 1:]:
        req_name = row[col_map["requirement"]] if "requirement" in col_map else None
        if not req_name or str(req_name).strip() == "":
            continue
        r = {
            "requirement": str(req_name).strip(),
            "description": str(row[col_map["description"]]).strip()
                            if "description" in col_map and row[col_map["description"]]
                            else "",
            "status": _normalise_status(str(row[col_map["status"]]).strip())
                      if "status" in col_map and row[col_map["status"]]
                      else "",
            "signal": str(row[col_map["signal"]]).strip()
                      if "signal" in col_map and row[col_map["signal"]]
                      else "",
        }
        # Carry domain column for single-sheet grouping
        if "domain" in col_map and row[col_map["domain"]]:
            r["_domain"] = str(row[col_map["domain"]]).strip()
        reqs.append(r)
    return reqs


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────
def parse_excel(path: str) -> list[dict]:
    """
    Parse an Excel workbook into the canonical requirements JSON structure.

    Returns:
        list of domain dicts:  [{name, description, reqs: [...]}]
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames

    # Strategy 1: multi-sheet (skip sheets that look like metadata)
    skip = {"summary", "overview", "metadata", "readme", "instructions", "cover"}
    domain_sheets = [s for s in sheets if s.strip().lower() not in skip]

    all_reqs = []
    for sheet_name in domain_sheets:
        ws = wb[sheet_name]
        reqs = _parse_sheet(ws)
        if reqs:
            all_reqs.append((sheet_name, reqs))
    wb.close()

    if not all_reqs:
        raise ValueError(f"No valid requirement rows found in {path}")

    # If only 1 sheet produced results, try grouping by _domain column
    if len(all_reqs) == 1:
        _, reqs = all_reqs[0]
        has_domain = any("_domain" in r for r in reqs)
        if has_domain:
            from collections import OrderedDict
            groups = OrderedDict()
            for r in reqs:
                dom = r.pop("_domain", "Uncategorized")
                groups.setdefault(dom, []).append(r)
            return [
                {
                    "name": dom,
                    "description": f"{len(rlist)} requirements",
                    "reqs": rlist,
                }
                for dom, rlist in groups.items()
            ]

    # Multi-sheet mode: each sheet = one domain
    result = []
    for i, (sheet_name, reqs) in enumerate(all_reqs, 1):
        # Clean _domain keys
        for r in reqs:
            r.pop("_domain", None)
        result.append({
            "name": f"Domain {i} of {len(all_reqs)} · {sheet_name}",
            "description": f"{len(reqs)} requirements",
            "reqs": reqs,
        })

    return result


def parse_excel_to_json(excel_path: str, output_path: str = None) -> str:
    """
    Parse Excel → JSON, optionally save to disk.
    Returns the JSON string.
    """
    data = parse_excel(excel_path)
    json_str = json.dumps(data, indent=2, ensure_ascii=False)
    if output_path:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(json_str)
    return json_str
